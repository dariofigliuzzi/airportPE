# ------------------------------------------------ #
# - SCRIPT PER GENERAZIONE GRAFICI A DISPERSIONE - #
# - SU FOGLI EXCEL CREATI DALLO SCRIPT loadExcel - #
# ------------------------------------------------ #
# - Dario Figliuzzi ------------------------------ #

import openpyxl
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

file = r'c:/Users/dario/git/airportPE/simulations/results/log2.xlsx'
wb = openpyxl.load_workbook(file)
worksheets = wb.sheetnames

# ---------------LOOK FOR EMPTY CELL-------------------------------------


def all_sheets(sheet_name):
    chart = ScatterChart()
    chart.title = "Scatter Chart"
    chart.style = 2
    last_index_landing, last_index_parking, last_index_takeoff = 0, 0, 0
    ws = wb[sheet_name]
    for i in range(2, ws.max_row+2):
        for j in range(8, 19, 5):
            if ws.cell(row=i, column=j).value is None:
                if j == 8 and last_index_landing == 0:
                    last_index_landing = i-1
                elif j == 13 and last_index_parking == 0:
                    last_index_parking = i-1
                elif j == 18 and last_index_takeoff == 0:
                    last_index_takeoff = i-1

    print(last_index_landing, last_index_parking, last_index_takeoff)

    # -------------------LANDING---------------------------------------------

    x_landing = Reference(ws, min_col=8, min_row=2, max_row=last_index_landing)
    y_landing = Reference(ws, min_col=9, min_row=2, max_row=last_index_landing)
    seriesLanding = Series(y_landing, x_landing, title="landing_queue")
    chart.series.append(seriesLanding)

    # -------------------TAKEOFF---------------------------------------------

    x_takeoff = Reference(ws, min_col=18, min_row=2, max_row=last_index_takeoff)
    y_takeoff = Reference(ws, min_col=19, min_row=2, max_row=last_index_takeoff)
    seriesTakeoff = Series(y_takeoff, x_takeoff, title="takeoff_queue")
    chart.series.append(seriesTakeoff)

    # -------------------PARKING---------------------------------------------

    x_parking = Reference(ws, min_col=13, min_row=2, max_row=last_index_parking)
    y_parking = Reference(ws, min_col=14, min_row=2, max_row=last_index_parking)
    seriesParking = Series(y_parking, x_parking, title="parking_queue")
    chart.series.append(seriesParking)

    # -----------------------------------------------------------------------

    ws.add_chart(chart, "A10")

    wb.save(file)


for i in worksheets:
    all_sheets(i)
