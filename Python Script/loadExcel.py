# --------------------------------------------------- #
# - SCRIPT PER CREAZIONE FOGLIO DI LAVORO EXCEL ----- #
# - PRELEVANDO I DATI DAI FILE RESULTS DI OMNET ----- #
# - AGGIORNAMENTO DELLO SCRIPT loadSheet IN SEGUITO - #
# - ALL'AGGIUNTA DI NUOVI VETTORI NEL FILE .VEC ----- #
# --------------------------------------------------- #
# - Dario Figliuzzi --------------------------------- #

import openpyxl
import re
from openpyxl.styles import PatternFill

path = r'c:/Users/dario/git/airportPE/simulations/results/'
txtPath = path + 'General-0-r'
excelPath = path + 'log2.xlsx'
iniPath = r'c:/Users/dario/git/airportPE/simulations/omnetpp.ini'
num_repeat = 0
wb = openpyxl.Workbook()
std = wb['Sheet']
wb.remove(std)

redFill = PatternFill(start_color='FCA5A5', end_color='FCA5A5', fill_type='solid')
blueFill = PatternFill(start_color='9CB0FF', end_color='9CB0FF', fill_type='solid')
yellowFill = PatternFill(start_color='FBFF62', end_color='FBFF62', fill_type='solid')
greenFill = PatternFill(start_color='98D79A', end_color='98D79A', fill_type='solid')

# ----------------LOOK FOR #REPETITION-------------------------------------------------

with open(iniPath, 'r') as file:
    filedata = file.readlines()

for i in range(len(filedata)):
    if 'repeat=' in filedata[i]:
        index = filedata[i].index('repeat=')
        num_repeat = int(filedata[i][index+7])

print(num_repeat)
file.close()

# ----------------------------PUT VECTORS IN THE SHEET---------------------------------


def load_data(file_number):

    # Ricavo indice riga di inizio vettori (diverso per ogni file)
    start_vec = 0
    with open(txtPath + str(file_number + 1) + '.vec', 'r') as f:
        fd = f.readlines()

    for i in range(len(fd)):
        if fd[i].startswith('0') and start_vec == 0:
            start_vec = i

    f.close()

    with open(txtPath + str(file_number + 1) + '.vec') as file:
        content = file.readlines()[start_vec:-1]

    content = [x.strip() for x in content]
    startColumnIndex = [1, 6, 11, 16]
    ws = wb.create_sheet('Sheet' + str(file_number+1))
    print(wb.worksheets)
    for i in startColumnIndex:
        if i == startColumnIndex[0]:
            ws.cell(row=1, column=i).value = "SKY"

        elif i == startColumnIndex[1]:
            ws.cell(row=1, column=i).value = "LANDING"

        elif i == startColumnIndex[2]:
            ws.cell(row=1, column=i).value = "PARKING"

        elif i == startColumnIndex[3]:
            ws.cell(row=1, column=i).value = "TAKEOFF"

        ws.cell(row=1, column=i+1).value = "EVENT ID"
        ws.cell(row=1, column=i+2).value = "TIME"
        if i != 1:
            ws.cell(row=1, column=i+3).value = "#inQUEUE"
        else:
            ws.cell(row=1, column=i+3).value = "PLANE ID"

    newcontent = [None] * len(content)
    index1, index2, index3, index4, index5, index6 = 0, 0, 0, 0, 0, 0

    for i in range(len(content)):
        line = content[i]
        newline = re.split(r'\t+', line.rstrip('\t'))

        if newline[0] == '1' and index1 == 0:
            index1 = i

        elif newline[0] == '2' and index2 == 0:
            index2 = i

        elif newline[0] == '3' and index3 == 0:
            index3 = i

        elif newline[0] == '4' and index4 == 0:
            index4 = i

        elif newline[0] == '5' and index5 == 0:
            index5 = i

        elif newline[0] == '6' and index6 == 0:
            index6 = i

        newcontent[i] = newline

    for i in range(len(newcontent)):
        n = newcontent[i]
        for j in range(len(n)):
            if i in range(0, index1):
                ws.cell(row=i + 2, column=j + startColumnIndex[0]).value = float(n[j])

            elif i in range(index1, index2):
                ws.cell(row=i - index1 + 2, column=j + startColumnIndex[1]).value = float(n[j])

            elif i in range(index3, index4):
                ws.cell(row=i - index3 + 2, column=j + startColumnIndex[2]).value = float(n[j])

            elif i in range(index5, index6):
                ws.cell(row=i - index5 + 2, column=j + startColumnIndex[3]).value = float(n[j])

    for i in range(0, index1+1):
        for j in range(startColumnIndex[0], startColumnIndex[0]+4):
            ws.cell(row=i + 1, column=startColumnIndex[0]).fill = redFill
            ws.cell(row=1, column=j).fill = redFill

    for i in range(0, index2 - index1 + 1):
        for j in range(startColumnIndex[1], startColumnIndex[1] + 4):
            ws.cell(row=i + 1, column=startColumnIndex[1]).fill = greenFill
            ws.cell(row=1, column=j).fill = greenFill

    for i in range(0, index4 - index3 + 1):
        for j in range(startColumnIndex[2], startColumnIndex[2] + 4):
            ws.cell(row=i + 1, column=startColumnIndex[2]).fill = yellowFill
            ws.cell(row=1, column=j).fill = yellowFill

    for i in range(0, index6 - index5 + 1):
        for j in range(startColumnIndex[3], startColumnIndex[3] + 4):
            ws.cell(row=i + 1, column=startColumnIndex[3]).fill = blueFill
            ws.cell(row=1, column=j).fill = blueFill

    file.close()
    wb.save(excelPath)

# --------------------------LOAD_DATA FOR EACH VEC FILE------------------------------


for a in range(0, num_repeat-1):
    load_data(a)
